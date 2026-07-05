"""组价成果导出（M2 §3.1 管线⑦）—— 会话状态 → 清单-定额-单价表（CSV / XLSX）。

定位：批量组价链路的**出口**。造价员最终交付在既有计价软件里完成（架构 §1.1：算术段只做
内部验证＋可导出），本模块把会话累积态整成可核对、可导入的表——逐行带 provenance
（编码来源/置信/确认方式），汇总区如实标注缺口（缺价件数、blocked 原因），**绝不补数**。

分层：``build_rows``（state.values → 表行，纯函数零依赖，本地可测）；``to_csv_bytes``
（标准库 csv，utf-8-sig BOM 防 Excel 中文乱码）；``to_xlsx_bytes``（延迟 import openpyxl，
服务器 ``uv add openpyxl``；未装则 ImportError 上抛由端点映射 501，CSV 恒可用）。
"""
from __future__ import annotations

import csv
import io
from typing import Any

# 明细表列头（清单-定额-单价表，对齐造价核对习惯；provenance 三列支撑审计）
ITEM_HEADERS = [
    "序号", "构件描述", "清单编码", "工程量", "定额子目",
    "人工费", "材料费", "机械费", "综合单价", "合价(不含税)",
    "状态", "编码来源", "置信度", "确认方式",
]


def _fmt(v: Any) -> Any:
    """空值 → 空串（表格里不出现 None 字样）；数值原样交 csv/xlsx 写。"""
    return "" if v is None else v


def _item_row(idx: int, item: dict[str, Any]) -> list[Any]:
    """单构件 → 明细行。缺口如实留空/标状态（missing_base / missing_quantity 等），不补数。

    code 两种形状都认：过闸钉值后是 lock_value ``{value, provenance, by}``（provenance 顶层）；
    未过闸是 ``{envelope, value, locked:False}``（provenance 在 envelope 里）——半程导出是常态
    （评审中途拉表核对），未确认件也要能看到候选来源与置信。
    """
    code = item.get("code") or {}
    prov = (code.get("provenance")
            or (code.get("envelope") or {}).get("provenance") or {})
    up = item.get("unit_price") or {}
    breakdown = up.get("breakdown") or {}
    status = up.get("status") or ("ok" if up.get("total_price") is not None else "")
    return [
        idx,
        _fmt(item.get("feature")),
        _fmt(code.get("value")),
        _fmt(item.get("quantity")),
        _fmt((item.get("quota") or {}).get("value")),
        _fmt(breakdown.get("人工费")),
        _fmt(breakdown.get("材料费")),
        _fmt(breakdown.get("施工机具使用费")),
        _fmt(up.get("unit_price")),
        _fmt(up.get("total_price")),
        _fmt(status),
        _fmt(prov.get("source_ref") or prov.get("source_type")),
        _fmt(prov.get("confidence")),
        ("人工确认" if code.get("by") == "user"
         else "自动采纳" if code.get("by") == "model"
         else "未确认" if code.get("value") else ""),
    ]


def build_rows(values: dict[str, Any]) -> dict[str, list[list[Any]]]:
    """会话状态 → ``{"items": 明细行, "summary": 汇总行}``（纯函数，导出格式共用的单一取数源）。

    参数：values —— ``session.get_state()['values']``（items / rates / params / rollup）。
    返回：明细行含表头；汇总区透传 rollup 的确定性结果（subtotal/措施/其他/规费/税前/税金/总价），
      rollup 缺失（会话未走到汇总）→ 汇总区只出说明行；blocked → 如实带 blocked_reason。
    """
    items = values.get("items") or []
    item_rows: list[list[Any]] = [list(ITEM_HEADERS)]
    for i, it in enumerate(items, 1):
        item_rows.append(_item_row(i, it))

    summary: list[list[Any]] = []
    rollup = values.get("rollup") or {}
    if rollup.get("blocked_reason"):
        summary.append(["会话状态", f"blocked：{rollup['blocked_reason']}"])
    elif rollup.get("total") is not None:
        summary += [
            ["分部分项合计", rollup.get("subtotal")],
            ["措施项目费", rollup.get("measure_fee")],
            ["其他项目费", rollup.get("other_fee")],
            ["规费", rollup.get("fee_levy")],
            ["税前造价", rollup.get("pre_tax_total")],
            ["税金", rollup.get("tax")],
            ["含税总造价", rollup.get("total")],
        ]
        missing = rollup.get("missing_unit_price_items")
        if missing:
            summary.append(["⚠ 缺综合单价构件数", missing])
    else:
        summary.append(["说明", "会话尚未完成汇总（未到总造价复核），仅导出已办明细"])

    rates = values.get("rates") or {}
    if rates:
        summary.append(["费率口径", f"管理费{_fmt(rates.get('management_fee_rate'))}% / "
                                    f"利润{_fmt(rates.get('profit_rate'))}% / "
                                    f"取费基数 {_fmt(rates.get('fee_base'))}"])
    summary.append(["口径声明", f"{values.get('region') or '深圳'}·{values.get('spec_version') or ''}"
                                "（本表仅供参考，不替代专业造价审核）"])
    return {"items": item_rows, "summary": summary}


def to_csv_bytes(rows: dict[str, list[list[Any]]]) -> bytes:
    """表行 → CSV 字节（utf-8-sig BOM：Excel 直开中文不乱码）。零第三方依赖，恒可用。"""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerows(rows["items"])
    w.writerow([])
    w.writerows(rows["summary"])
    return buf.getvalue().encode("utf-8-sig")


def to_xlsx_bytes(rows: dict[str, list[list[Any]]]) -> bytes:
    """表行 → XLSX 字节（openpyxl，服务器 ``uv add openpyxl``；未装 ImportError 上抛→端点映射 501）。

    两个 sheet：清单明细 / 汇总。样式极简（表头加粗+列宽），核对与导入优先、不做报表美化
    （正式报表在计价软件出，架构 §1.1）。
    """
    from openpyxl import Workbook  # 延迟 import：CSV 路径不吃此依赖
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "清单明细"
    for row in rows["items"]:
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    widths = [5, 40, 14, 9, 12, 9, 9, 9, 10, 12, 10, 30, 8, 9]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws2 = wb.create_sheet("汇总")
    for row in rows["summary"]:
        ws2.append(row)
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 60

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
