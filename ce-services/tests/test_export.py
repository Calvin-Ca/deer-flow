#!/usr/bin/env python3
"""组价成果导出回归（M2 管线⑦）：build_rows 纯函数 + CSV 字节（零依赖本地全测）；
xlsx 分支本地无 openpyxl 自动跳过（服务器 pytest 全跑）。双模式同既有约定。

样例 state 形状对齐真实链路：lock_value 结构的 code、compute_unit_price 结构的 unit_price、
rollup_cost+single_works 结构的 rollup——键名改动会在此处红（导出层的契约哨兵）。
"""
from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from cost.export import ITEM_HEADERS, build_rows, to_csv_bytes  # noqa: E402

_VALUES = {
    "region": "深圳", "spec_version": "2024",
    "items": [
        {  # 正常件：人工确认的码 + 完整算价
            "feature": "C30现浇钢筋混凝土独立基础", "quantity": 120.0,
            "code": {"value": "010501003001", "locked": True, "by": "user",
                     "provenance": {"source_type": "spec_clause", "source_ref": "GB50854-2024 附录E",
                                    "confidence": 0.72}},
            "quota": {"value": "A1-15", "locked": True, "by": "model", "provenance": {}},
            "unit_price": {"unit_price": 365.0, "total_price": 43800.0,
                           "breakdown": {"人工费": 100.0, "材料费": 200.0, "施工机具使用费": 50.0}},
        },
        {  # 坏件：无定额基价（诚实标 missing_base，行留空不补数）
            "feature": "神秘构件",
            "code": {"value": "019999999999", "locked": True, "by": "model",
                     "provenance": {"source_type": "spec_clause", "confidence": 0.9}},
            "unit_price": {"status": "missing_base"},
        },
    ],
    "rates": {"management_fee_rate": 10.0, "profit_rate": 5.0, "fee_base": "labor"},
    "rollup": {"subtotal": 43800.0, "measure_fee": 0.0, "other_fee": 0.0, "fee_levy": 0.0,
               "pre_tax_total": 43800.0, "tax": 3942.0, "total": 47742.0,
               "missing_unit_price_items": 1, "single_works": []},
}


def test_build_rows_items():
    rows = build_rows(_VALUES)
    assert rows["items"][0] == ITEM_HEADERS
    good = rows["items"][1]
    assert good[0] == 1 and good[2] == "010501003001" and good[3] == 120.0
    assert good[8] == 365.0 and good[9] == 43800.0
    assert good[11] == "GB50854-2024 附录E" and good[12] == 0.72 and good[13] == "人工确认"
    bad = rows["items"][2]
    assert bad[10] == "missing_base" and bad[9] == "" and bad[13] == "自动采纳"  # 缺口留空不补数


def test_build_rows_summary_with_rollup():
    summary = build_rows(_VALUES)["summary"]
    kv = {r[0]: r[1] for r in summary if len(r) == 2}
    assert kv["含税总造价"] == 47742.0 and kv["税金"] == 3942.0
    assert kv["⚠ 缺综合单价构件数"] == 1  # 缺口如实进汇总
    assert "管理费10.0%" in kv["费率口径"]
    assert "深圳·2024" in kv["口径声明"]


def test_build_rows_unconfirmed_code_shape():
    """未过闸形状（真会话半程导出实测缺口）：provenance 在 envelope 里也要读到，确认方式标「未确认」。"""
    values = {"items": [{
        "feature": "C30柱",
        "code": {"value": "010502001", "locked": False,
                 "envelope": {"provenance": {"source_ref": "GB50854-2013 附录E", "confidence": 0.8}}},
    }]}
    row = build_rows(values)["items"][1]
    assert row[2] == "010502001" and row[11] == "GB50854-2013 附录E" and row[12] == 0.8
    assert row[13] == "未确认"


def test_build_rows_blocked_and_unfinished():
    blocked = build_rows({"items": [{"feature": "x"}],
                          "rollup": {"blocked_reason": "无法组价到总价（缺定额映射）"}})
    assert any("blocked" in str(r[1]) for r in blocked["summary"] if len(r) == 2)
    unfinished = build_rows({"items": [{"feature": "x"}]})
    assert any("尚未完成汇总" in str(r[1]) for r in unfinished["summary"] if len(r) == 2)


def test_csv_bytes_bom_and_content():
    data = to_csv_bytes(build_rows(_VALUES))
    assert data.startswith(b"\xef\xbb\xbf")  # utf-8-sig BOM：Excel 直开中文不乱码
    text = data.decode("utf-8-sig")
    assert "010501003001" in text and "含税总造价" in text and "missing_base" in text


def test_xlsx_bytes_two_sheets():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        if __name__ != "__main__":
            import pytest
            pytest.skip("openpyxl 未装（本地）——xlsx 分支服务器验")
        print("  (skip xlsx：本地无 openpyxl)")
        return
    import io

    from cost.export import to_xlsx_bytes
    data = to_xlsx_bytes(build_rows(_VALUES))
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["清单明细", "汇总"]
    assert wb["清单明细"].cell(row=2, column=3).value == "010501003001"


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass
    failed = 0
    for _name in sorted(k for k in dir() if k.startswith("test_")):
        try:
            globals()[_name]()
            print(f"✓ {_name}")
        except Exception as exc:  # noqa: BLE001
            failed += 1
            print(f"✗ {_name}  {type(exc).__name__}: {exc}")
    print(f"\n导出回归：{'全绿' if not failed else f'{failed} 败'}")
    sys.exit(1 if failed else 0)
