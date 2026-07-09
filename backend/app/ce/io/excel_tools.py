"""工程量清单（BOQ）Excel 的确定性解析工具。

造价任务大多从一张工程量清单表（广联达 GCL / 招标清单导出的 .xlsx）开始，
而通用的 markitdown 文档转换会丢掉合并单元格与列结构，拿不到结构化行。
本模块用 openpyxl 做确定性解析：按表头关键词定位表头行、映射列，再逐行抽取
「项目编码 / 项目名称 / 项目特征 / 计量单位 / 工程量（及可选的综合单价/合价）」。

分层：``parse_boq_workbook`` 是纯函数（输入 xlsx 字节，无沙箱依赖，可直接单测）；
``parse_boq_excel`` 是 ``@tool`` 包装，只负责从沙箱把上传文件下载成字节。
"""

from __future__ import annotations

import io
from typing import Any

from langchain.tools import tool

from deerflow.sandbox.tools import ensure_sandbox_initialized
from deerflow.tools.types import Runtime

# 单个列语义 → 表头可能出现的关键词（去空白后做子串匹配）。
# 顺序即优先级：一行里同一语义命中多列时，取第一个命中的列。
_COLUMN_KEYWORDS: dict[str, tuple[str, ...]] = {
    "code": ("项目编码", "清单编码", "项目编号", "编码"),
    "name": ("项目名称", "清单名称", "名称"),
    "feature": ("项目特征", "特征描述", "项目特征描述", "特征"),
    "unit": ("计量单位", "单位"),
    "quantity": ("工程量", "工程数量", "数量"),
    "unit_price": ("综合单价", "单价"),
    "total_price": ("综合合价", "合价", "金额"),
}

# 判定一行为「有效表头」所需的最少语义：必须能定位编码 + 名称 + （单位或工程量）。
_REQUIRED_ANY_QUANTITATIVE = ("unit", "quantity")

_MAX_HEADER_SCAN_ROWS = 30  # 表头之上通常只有标题/工程名称等少数行
_MAX_ITEMS = 5000  # 抽取行数上限，超出则截断并标记
_BLANK_STREAK_STOP = 15  # 连续空行达到该值即认为清单结束


def _norm(value: Any) -> str:
    """把单元格值归一成用于关键词匹配的紧凑字符串。

    功能：去除首尾及内部空白/换行，便于「项 目 编 码」这类带空格表头的匹配。
    参数：value —— 任意单元格原始值（可能是 None / 数字 / 字符串）。
    返回：紧凑后的字符串；None 归一为空串。
    """
    if value is None:
        return ""
    return "".join(str(value).split())


def _detect_header(rows: list[list[Any]]) -> tuple[int | None, dict[str, int]]:
    """在前若干行里定位清单表头行，并建立「语义 → 列下标」映射。

    功能：对每一候选行按命中的列语义数量打分，取分数最高且满足最小语义要求
        （编码+名称+（单位或工程量））的行作为表头；并列时取最靠上的一行。
    参数：rows —— 全表二维单元格值（外层行、内层列）。
    返回：(表头行下标或 None, {语义: 列下标})；未识别到合法表头时行下标为 None。
    """
    best_idx: int | None = None
    best_mapping: dict[str, int] = {}
    best_score = 0

    for idx, row in enumerate(rows[:_MAX_HEADER_SCAN_ROWS]):
        mapping: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            text = _norm(cell)
            if not text:
                continue
            for semantic, keywords in _COLUMN_KEYWORDS.items():
                if semantic in mapping:
                    continue  # 同一语义只取第一个命中的列
                if any(kw in text for kw in keywords):
                    mapping[semantic] = col_idx
                    break
        has_required = "code" in mapping and "name" in mapping and any(s in mapping for s in _REQUIRED_ANY_QUANTITATIVE)
        score = len(mapping)
        if has_required and score > best_score:
            best_idx, best_mapping, best_score = idx, mapping, score

    return best_idx, best_mapping


def _to_number(value: Any) -> float | None:
    """把工程量/价格单元格解析成数值。

    功能：直接取数字；字符串则去掉千分位逗号与首尾空白后转 float；不可解析返回 None。
    参数：value —— 单元格原始值。
    返回：float 或 None。
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _to_code(value: Any) -> str:
    """把项目编码单元格归一成字符串，保留前导零。

    功能：字符串直接去空白；若被存成数字（如 010502001001 被读成 float），
        还原为不带小数点的整数字符串，避免丢失编码位。
    参数：value —— 编码单元格原始值。
    返回：编码字符串；空值返回空串。
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def parse_boq_workbook(data: bytes, sheet: str | None = None) -> dict[str, Any]:
    """解析工程量清单 Excel 字节流为结构化行（纯函数，无沙箱依赖）。

    功能：加载 xlsx、选定工作表、按表头关键词定位表头并映射列，逐行抽取清单项。
        不做任何造价推断，只忠实抽取表内已有内容；识别不到合法表头时如实报错并附预览。
    参数：
        data: xlsx 文件的二进制内容。
        sheet: 可选工作表名；缺省时自动选第一个能识别出清单表头的工作表。
    返回：
        成功 —— {"status": "ok", "sheet", "sheet_names", "header_row"(1基),
            "columns"(语义→列名占位), "row_count", "truncated", "items"[...], "warnings"[...]}；
            每个 item 含 row(1基)、code、name、feature、unit、quantity、unit_price、total_price、is_bill。
        失败 —— {"status": "error", "error", "sheet_names", "preview"(前若干行)}。
    """
    try:
        import openpyxl
    except ImportError:  # pragma: no cover - 依赖缺失时给出可执行的提示
        return {"status": "error", "error": "openpyxl 未安装，请在 backend 执行 `uv add openpyxl`。"}

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as e:  # openpyxl 会对损坏/非 xlsx 抛多种异常
        return {"status": "error", "error": f"无法打开为 Excel 工作簿：{e}"}

    try:
        sheet_names = list(wb.sheetnames)
        if sheet is not None and sheet not in sheet_names:
            return {
                "status": "error",
                "error": f"工作表 '{sheet}' 不存在。",
                "sheet_names": sheet_names,
            }

        candidate_names = [sheet] if sheet is not None else sheet_names
        first_preview: list[list[Any]] = []
        for name in candidate_names:
            ws = wb[name]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            if not first_preview:
                first_preview = [[_norm(c) for c in r] for r in rows[:10]]
            header_idx, mapping = _detect_header(rows)
            if header_idx is None:
                continue
            return _extract(rows, header_idx, mapping, name, sheet_names)

        return {
            "status": "error",
            "error": "未能在任一工作表中识别出工程量清单表头（需包含「项目编码/项目名称/计量单位或工程量」等列）。请确认这是一份工程量清单表。",
            "sheet_names": sheet_names,
            "preview": first_preview,
        }
    finally:
        wb.close()


def _extract(
    rows: list[list[Any]],
    header_idx: int,
    mapping: dict[str, int],
    sheet_name: str,
    sheet_names: list[str],
) -> dict[str, Any]:
    """按已定位的表头行与列映射抽取数据行。

    功能：从表头下一行起逐行读取，遇到连续空行超阈值即停止；忠实回填各语义列的值。
    参数：rows 全表值、header_idx 表头行下标、mapping 语义→列下标、sheet_name、sheet_names。
    返回：成功结构（见 ``parse_boq_workbook`` 的成功分支）。
    """

    def cell(row: list[Any], semantic: str) -> Any:
        col = mapping.get(semantic)
        if col is None or col >= len(row):
            return None
        return row[col]

    items: list[dict[str, Any]] = []
    warnings: list[str] = []
    blank_streak = 0
    truncated = False

    for offset, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        if not any(_norm(c) for c in row):
            blank_streak += 1
            if blank_streak >= _BLANK_STREAK_STOP:
                break
            continue
        blank_streak = 0

        code = _to_code(cell(row, "code"))
        name = str(cell(row, "name")).strip() if cell(row, "name") is not None else ""
        feature = str(cell(row, "feature")).strip() if cell(row, "feature") is not None else ""
        unit = str(cell(row, "unit")).strip() if cell(row, "unit") is not None else ""
        quantity = _to_number(cell(row, "quantity"))
        unit_price = _to_number(cell(row, "unit_price"))
        total_price = _to_number(cell(row, "total_price"))

        # 全空语义行（可能是分隔/小计的纯格式行）直接跳过，不计入清单项。
        if not (code or name or feature or unit or quantity is not None):
            continue

        items.append(
            {
                "row": offset,
                "code": code,
                "name": name,
                "feature": feature,
                "unit": unit,
                "quantity": quantity,
                "unit_price": unit_price,
                "total_price": total_price,
                "is_bill": bool(code),  # 有编码=清单子目；无编码=分部/标题行
            }
        )
        if len(items) >= _MAX_ITEMS:
            truncated = True
            warnings.append(f"清单行数超过 {_MAX_ITEMS}，已截断。")
            break

    if "quantity" in mapping and items and all(it["quantity"] is None for it in items):
        warnings.append("已定位工程量列，但所有行工程量为空——文件可能未在 Excel 中打开过（公式无缓存值），或工程量确未填写。")

    columns = {semantic: _column_letter(col) for semantic, col in mapping.items()}
    return {
        "status": "ok",
        "sheet": sheet_name,
        "sheet_names": sheet_names,
        "header_row": header_idx + 1,
        "columns": columns,
        "row_count": len(items),
        "truncated": truncated,
        "items": items,
        "warnings": warnings,
    }


def _column_letter(col_idx: int) -> str:
    """把 0 基列下标转成 Excel 列字母（0→A，26→AA），仅用于回显定位。"""
    result = ""
    n = col_idx + 1
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(ord("A") + rem) + result
    return result


@tool("parse_boq_excel", parse_docstring=True)
def parse_boq_excel(runtime: Runtime, path: str) -> dict[str, Any]:
    """解析用户上传的工程量清单 Excel（.xlsx），返回结构化清单行。

    用于把一份工程量清单表（如广联达导出、招标清单）读成结构化数据，供后续组价/核对使用。
    只忠实抽取表内已有内容（项目编码/名称/特征/单位/工程量等），不做任何造价推断或补全；
    识别不到清单表头时会返回 error 并附前几行预览，请据此向用户澄清，切勿臆造清单内容。

    Args:
        path: 上传文件的绝对路径，须位于 /mnt/user-data/ 下（通常是 /mnt/user-data/uploads/xxx.xlsx）。
    """
    try:
        sandbox = ensure_sandbox_initialized(runtime)
        data = sandbox.download_file(path)
    except FileNotFoundError:
        return {"status": "error", "error": f"文件不存在：{path}"}
    except PermissionError:
        return {"status": "error", "error": f"路径不在允许目录内（须位于 /mnt/user-data/ 下）：{path}"}
    except Exception as e:
        return {"status": "error", "error": f"读取文件失败：{e}", "path": path}
    return parse_boq_workbook(data)


parse_boq_excel_tool = parse_boq_excel

__all__ = ["parse_boq_excel", "parse_boq_excel_tool", "parse_boq_workbook"]
